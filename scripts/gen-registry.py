#!/usr/bin/env python3
"""src/data/registry.ts генерацияси — ССВ идоралараро алмашинув реестридан.

Фойдаланиш:
    python3 scripts/gen-registry.py "<xlsx йўли>"

xlsx формати (1-варақ): 4–64 қаторлар, B–H устунлар:
№ | ССВ тизими | идора тизими | алмашинадиган маълумотлар | ҳуқуқий асос | ҳужжат | ҳолат.

Ҳар бир қаторнинг идораси AGENCY_OF харитасида (№ → идора id);
реестр янгиланганда янги қаторлар учун харитани тўлдиринг.
"""
import json
import re
import sys

import openpyxl

# реестр қаторининг идораси (№ → id) — 30.03.2026 ҳолати
AGENCY_OF = {
    **dict.fromkeys([8, 9, 10, 11, 12, 13, 14, 16, 17, 18, 35, 36, 37, 44, 45, 46, 47], 'social'),
    **dict.fromkeys([1, 2, 3, 4, 25], 'etrm'),
    **dict.fromkeys([5, 6, 7], 'insurance'),
    **dict.fromkeys([21, 22, 23, 24], 'hisob'),
    **dict.fromkeys([26, 27, 38, 50, 54, 55], 'soliq'),
    **dict.fromkeys([29, 30, 31], 'iiv'),
    **dict.fromkeys([34, 48, 49], 'moliya'),
    **dict.fromkeys([19, 20, 32, 33], 'mudofaa'),
    **dict.fromkeys([28, 52], 'adliya'),
    15: 'yoshlar', 39: 'maktab', 40: 'oila', 51: 'tech', 53: 'pharm',
    **dict.fromkeys([41, 42, 43, 56, 57, 58, 59, 60, 61], 'rh'),
}

# id → (тўлиқ ном, харитадаги ёрлиқ, доира ичидаги белги)
# Ёрлиқ — расмий номнинг ўзи (қисқартириб бузилмайди); фақат ЭТРМ —
# расмий аббревиатура (тўлиқ номи ҳалқа ёрлиғи учун жуда узун).
AG = {
    'social': ('Ижтимоий ҳимоя миллий агентлиги', 'Ижтимоий ҳимоя миллий агентлиги', 'ИҲ'),
    'etrm': ('Электрон технологияларни ривожлантириш маркази', 'ЭТРМ', 'ЭТРМ'),
    'insurance': ('Давлат тиббий суғурта жамғармаси', 'Давлат тиббий суғурта жамғармаси', 'ДТСЖ'),
    'hisob': ('Ҳисоб палатаси', 'Ҳисоб палатаси', 'ҲП'),
    'soliq': ('Солиқ қўмитаси', 'Солиқ қўмитаси', 'СҚ'),
    'iiv': ('Ички ишлар вазирлиги', 'Ички ишлар вазирлиги', 'ИИВ'),
    'moliya': ('Иқтисодиёт ва молия вазирлиги', 'Иқтисодиёт ва молия вазирлиги', 'ИМВ'),
    'mudofaa': ('Мудофаа вазирлиги', 'Мудофаа вазирлиги', 'МВ'),
    'adliya': ('Адлия вазирлиги', 'Адлия вазирлиги', 'АВ'),
    'yoshlar': ('Ёшлар ишлари агентлиги', 'Ёшлар ишлари агентлиги', 'ЁИА'),
    'maktab': ('Мактаб ва мактабгача таълим вазирлиги', 'Мактаб ва мактабгача таълим вазирлиги', 'МТВ'),
    'oila': ('Оила ва хотин-қизлар қўмитаси', 'Оила ва хотин-қизлар қўмитаси', 'ОХҚ'),
    'tech': ('Техник тартибга солиш агентлиги', 'Техник тартибга солиш агентлиги', 'ТТС'),
    'pharm': ('Фармацевтика ривожлантириш агентлиги', 'Фармацевтика', 'ФРА'),
    'oits': ('ОИТСга қарши курашиш маркази', 'ОИТС маркази', 'ОИТС'),
    'rh': ('«Рақамли ҳукумат» платформаси', '«Рақамли ҳукумат» платформаси', 'РҲ'),
}

# ССВ тизими: (тўлиқ ном ичидаги калит, қисқа ном, белги, иконка калити)
MOH_SHORT = [
    ('DMED', 'DMED — Ягона тиббиёт АТ', 'DMED', 'dmed'),
    ('MIS2', 'MIS2 — Тиббиёт АТ', 'MIS2', 'mis2'),
    ('narko-psix', 'narko-psix.uz', 'NPX', 'npx'),
    ('ОИТСга қарши', 'HIV-ES — ОИТС маркази', 'HIV', 'hiv'),
    ('Туғилиш ва ўлимни', 'Туғилиш/ўлим қайди АТ', 'Т/Ў', 'birth'),
    ('Рақамли соғлиқ', 'Рақамли соғлиқ платформаси', 'РСП', 'rsp'),
    ('Medrefer', 'Medrefer — суғурта', 'MDR', 'mdr'),
    ('Cancer', 'Cancer-registr', 'CANC', 'canc'),
    ('Донор', 'Донор ҳисоби — Қон хизмати', 'ДОНОР', 'donor'),
    ('ЖиҳозМед', 'ЖиҳозМед АТ', 'ЖМ', 'jm'),
    ('МедДата', 'МедДата — тез ёрдам', '103', 'amb'),
    ('Автохўжалик', 'Автохўжалик АТ', 'АВТО', 'avto'),
]

# №1–43 — ССВ идораларга маълумот ТАҚДИМ ЭТАДИ (узатиш),
# №44–61 — идоралар ССВга маълумот ТАҚДИМ ЭТАДИ (қабул қилиш).
OUTGOING_MAX = 43

# Харитадан бутунлай чиқарилган идоралар (фойдаланувчи талаби)
EXCLUDE_AGENCIES = {'pharm'}
# Чиқарилган реестр қаторлари (фойдаланувчи талаби):
# №54 — РСП ⇄ Рақамли ҳукумат (Солиқ); №26 — дори маркировкаси (Солиқ);
# №19, №32, №33 — Мудофаа: улар ўрнига №20 нинг 9 банди алоҳида кўрсатилади
EXCLUDE_ROWS = {19, 26, 32, 33, 54}

# Реестрда «Жараёнда» деб турса-да, амалда ишлаётган қаторлар (фойдаланувчи тасдиқлади):
# №56 — вояга етмаганлар, №60 — ҳомиладор аёллар ва бола парвариши,
# №21 — Ҳисоб палатаси: рўйхатга олинган беморлар ва муассасалар маълумоти,
# №25 — ОИВ инфекцияси бўлган шахслар (ЭТРМ)
FORCE_LIVE = {21, 25, 56, 60}

# Реестрда «Мавжуд» турса-да, амалда ҳали режада (фойдаланувчи тасдиқлади):
# №55 — онлайн назорат касса чеклари (Солиқ қўмитаси)
FORCE_PLAN = {55}

# Икки томонлама алмашинувлар (фойдаланувчи изоҳи):
# № → {'in': идора ССВга берадигани, 'out': ССВ берадиганига қўшимча банд}
BIDIR = {
    15: {
        'in': '30 ёшгача бўлган ёшлар тўғрисидаги маълумотлар',
        'out': '30 ёшгача бўлган ёшлар тўғрисидаги маълумотлар',
    },
}

# Реестр қаторларидан ташқари қўшимча йўналишлар (фойдаланувчи талаби):
# идора id → [{name, moh_key (MOH_SHORT калити), dir, live}]
EXTRA_LEAVES = {
    'adliya': [
        {
            'name': 'Туғилиш ва ўлим маълумотномасини олиш',
            'moh_key': 'Туғилиш ва ўлимни',
            'dir': 'out',
            'live': True,
        },
    ],
    # ВМнинг 2026 йил 29 январдаги 35-сон қарори бўйича режадаги йўналишлар
    'social': [
        {
            'name': 'Йилда бир марта тўлиқ тиббий кўрикдан ўтказиш',
            'moh_key': 'DMED',
            'dir': 'out',
            'live': False,
            'basis': 'Ўзбекистон Республикаси Вазирлар Маҳкамасининг 2026 йил 29 январдаги 35-сон қарори',
        },
        {
            'name': 'Қандли диабетнинг иккинчи тури билан касалланган ҳамда эҳтиёжманд беморларни метформин дори воситаси билан бепул таъминлаш',
            'moh_key': 'DMED',
            'dir': 'out',
            'live': False,
            'basis': 'Ўзбекистон Республикаси Вазирлар Маҳкамасининг 2026 йил 29 январдаги 35-сон қарори',
        },
        {
            'name': 'Аллергик ва иммунологик касалликларга чалинган беморларни ингаляцион глюкокортикостероид дори воситалар билан бепул таъминлаш',
            'moh_key': 'DMED',
            'dir': 'out',
            'live': False,
            'basis': 'Ўзбекистон Республикаси Вазирлар Маҳкамасининг 2026 йил 29 январдаги 35-сон қарори',
        },
    ],
    'oila': [
        {
            'name': 'Онкоскрининг тўғрисида хабардор қилиш',
            'moh_key': 'DMED',
            'dir': 'out',
            'live': False,
        },
        {
            'name': 'Психотроп воситаларга ружу қўйган ёки мойиллиги бор хотин-қизлар тўғрисида хабардор қилиш',
            'moh_key': 'DMED',
            'dir': 'out',
            'live': False,
        },
        {
            'name': 'Спиртли ичимликларга ружу қўйган ёки мойиллиги бор хотин-қизлар тўғрисида хабардор қилиш',
            'moh_key': 'DMED',
            'dir': 'out',
            'live': False,
        },
        {
            'name': 'Сурункали касалликка чалинган хотин-қизлар тўғрисида хабардор қилиш',
            'moh_key': 'DMED',
            'dir': 'out',
            'live': False,
        },
    ],
    # №14 ижтимоий ҳимояда ҳам қолади; фойдаланувчи талаби билан Ёшлар
    # агентлигида алоҳида йўналиш сифатида ҳам кўрсатилади.
    'yoshlar': [
        {
            'name': 'Йўлланма (ордер) шаклланганлиги тўғрисидаги маълумот (бунда Агентликнинг коллегиал қарори маълумотини интегратсия қилган ҳолда шакллантирилган бўлиши лозим)',
            'moh_key': 'MIS2',
            'dir': 'out',
            'live': True,
            'registry_no': 14,
            'count_in_total': False,
            'basis': 'Ўзбекистон Республика Вазирлар Маҳкамасининг 2025 йил 5 майдаги 294-сон қарори',
        },
    ],
}

# Вергул бўйича ажратилмайдиган қаторлар: матн — яхлит бир гап
# (№48 — «...шахслар, шунингдек, уларга тенглаштирилган шахслар»,
#  №49 — пенсионерларнинг хизмат жойлари санамаси,
#  №58 — «Солиқ кодексига мувофиқ, ...» кириш иборали гап)
NO_COMMA_SPLIT = {48, 49, 58}

# Кўп бандли қаторлар АВТОМАТИК алоҳида тугунларга бўлинади (ҳар банд —
# алоҳида барг). Яхлит қоладиганлар — NO_SPLIT_ROWS (№15 — икки томонлама
# карточка). SPLIT_ROWS — банд матнлари қўлда аниқлаштирилган қаторлар.
NO_SPLIT_ROWS = {15}

SPLIT_ROWS = {
    30: [
        ['Беморларнинг тиббий маълумотлари'],
        ['Бириктирилган поликлиникаси тўғрисида маълумот'],
        ['Қўйилган диагнозлари тўғрисида маълумот'],
        ['Касаллик тарихи тўғрисида маълумот'],
        ['Ҳомиладорлик ҳисобида турганли тўғрисида маълумот'],
    ],
    # №43 — банд матни фойдаланувчи талабига кўра қисқартирилган
    43: [
        ['Шахснинг иш жойлари тўғрисидаги маълумот олиш'],
    ],
    # Мудофаа: №20 нинг ҳар бир банди — алоҳида тугун (9 та)
    20: [
        ['Антропометрия (бўй, вазн, кўкрак қафаси атрофи ва бошқалар)'],
        ['Қоннинг тўлиқ таҳлили'],
        ['Нажаснинг тўлиқ таҳлили'],
        ['Сийдикнинг тўлиқ таҳлили'],
        ['Оилавий клиника ва оилавий шифокорга рўйхатдан ўтиш'],
        ['Амбулатория шароитида рўйхатдан ўтиш'],
        ['Текширувлар'],
        ['Инструментал текширувлар (ЭКГ, ультратовуш текшируви, ФГДС)'],
        ['Эмлашлар'],
    ],
}

YEAR_RE = re.compile(r'(20\d\d)\s*йил')


def clean(s):
    return re.sub(r'\s+', ' ', str(s or '')).strip()


def moh_short(full):
    for key, short, mark, icon in MOH_SHORT:
        if key in full:
            return short, mark, icon
    return full[:24], re.sub(r'[^\w]', '', full)[:4].upper(), None


def doc_label(doc):
    low = doc.lower()
    if 'ариза' in low:
        return 'Ариза берилган'
    if 'чиқилмоқда' in low:
        return 'Ишлаб чиқилмоқда'
    if 'чиқилган' in low or 'тасдиқланган' in low:
        return 'Тасдиқланган'
    return 'Ишлаб чиқилмоқда'


def comma_split(x):
    """Ҳар бир юқори даражадаги вергул — янги банд (вергулдан кейин бўшлиқ
    бўлса); қавс ичидаги вергуллар ҳисобга олинмайди («(ЭКГ, УТТ, ФГДС)»)."""
    parts, depth, start = [], 0, 0
    for i, ch in enumerate(x):
        if ch == '(':
            depth += 1
        elif ch == ')':
            depth = max(0, depth - 1)
        elif ch == ',' and depth == 0:
            j = i + 1
            while j < len(x) and x[j] == ' ':
                j += 1
            if j > i + 1 and j < len(x):
                parts.append(x[start:i])
                start = j
    parts.append(x[start:])
    return parts


def split_data(raw, comma=True):
    # аввал «;» ва янги қатор бўйича, сўнг (истисно бўлмаса) вергул бўйича
    items = [clean(x) for x in re.split(r'[;\n]+', str(raw or ''))]
    if comma:
        items = [p for x in items for p in comma_split(x)]
    # ячейкадаги «1.», «1.1.», «2)» каби рақамлаш ва четки қўштирноқларни оламиз
    items = [re.sub(r'^\d+(?:\.\d+)*\s*[.)]\s*', '', x).strip('"“” ') for x in items]
    return [x.rstrip('.').strip() for x in items if x and len(x) > 2]


def trunc(s, n):
    return s if len(s) <= n else s[: n - 1].rstrip() + '…'


def main(path):
    ws = openpyxl.load_workbook(path, data_only=True).worksheets[0]
    agency_children = {k: [] for k in AG}
    total = 0
    live_rows = 0
    out_rows = 0
    via_rows = 0
    acts = set()
    for r in range(4, 65):
        no = ws.cell(r, 2).value
        if no is None:
            continue
        no = int(no)
        if no in EXCLUDE_ROWS or AGENCY_OF[no] in EXCLUDE_AGENCIES:
            continue
        total += 1
        moh = clean(ws.cell(r, 3).value)
        partner = clean(ws.cell(r, 4).value)
        items = split_data(ws.cell(r, 5).value, comma=no not in NO_COMMA_SPLIT) or [clean(ws.cell(r, 5).value)]
        basis = clean(ws.cell(r, 6).value)
        doc = clean(ws.cell(r, 7).value)
        status = clean(ws.cell(r, 8).value)
        ag = AGENCY_OF[no]
        # Идора тизими устунида «Рақамли ҳукумат» платформаси кўрсатилган —
        # алмашинув тўғридан-тўғри эмас, платформа орқали (РҲнинг ўз
        # қаторлари бундан мустасно: улар платформанинг ўзи билан алмашинув)
        via = 'рақамли ҳукумат' in partner.lower() and ag != 'rh'
        short, mark, sys_icon = moh_short(moh)
        live = (status == 'Мавжуд' or no in FORCE_LIVE) and no not in FORCE_PLAN
        outgoing = no <= OUTGOING_MAX
        live_rows += 1 if live else 0
        out_rows += 1 if outgoing else 0
        via_rows += 1 if via else 0
        if basis:
            acts.add(basis)
        year = (YEAR_RE.search(basis) or [None, None])[1]
        both = no in BIDIR
        moh_step = {'title': short, 'sub': 'ССВ тизими', 'icon': sys_icon}
        ag_step = {'title': AG[ag][1], 'sub': 'ҳамкор идора', 'icon': ag}
        via_step = {'title': '«Рақамли ҳукумат»', 'sub': 'интеграция платформаси', 'icon': 'rh'}
        kpis = [
            {'value': 'ССВ ⇄ Идора' if both else 'ССВ → Идора' if outgoing else 'Идора → ССВ',
             'label': 'маълумот йўналиши'},
            {'value': 'Платформа орқали' if via else 'Тўғридан-тўғри', 'label': 'уланиш тури'},
            {'value': 'Мавжуд' if live else 'Режада', 'label': 'алмашинув ҳолати'},
            {'value': doc_label(doc), 'label': 'ҳужжат ҳолати'},
        ]
        if not live:
            kpis.append({'value': '2026 йил охири', 'label': 'режа муддати'})
        if year:
            kpis.append({'value': year, 'label': 'асос йили'})
        if both and BIDIR[no].get('out'):
            items.insert(0, BIDIR[no]['out'])
        if both:
            arrow_desc = f'{short} ⇄ {AG[ag][0]}. Икки томонлама алмашинув.'
        else:
            arrow_desc = (f'{short} → {AG[ag][0]}. ССВ маълумот тақдим этади.' if outgoing
                          else f'{AG[ag][0]} → {short}. Идора ССВга маълумот тақдим этади.')
        if via:
            arrow_desc += ' Алмашинув «Рақамли ҳукумат» платформаси орқали.'
        if not live:
            arrow_desc += ' 2026 йил охирига режалаштирилган.'
        # SPLIT_ROWS — қўлда белгиланган гуруҳлар; акс ҳолда кўп бандли қатор
        # автоматик банд-банд бўлинади (NO_SPLIT_ROWS дан ташқари)
        if no in SPLIT_ROWS:
            groups = SPLIT_ROWS[no]
        elif len(items) > 1 and no not in NO_SPLIT_ROWS:
            groups = [[it] for it in items]
        else:
            groups = [items]
        # тугун номи ва банд матни бош ҳарф билан бошлансин
        groups = [[(g[0].upper() + g[1:]) if g else g for g in grp] for grp in groups]
        chain = [moh_step, ag_step] if outgoing or both else [ag_step, moh_step]
        if via:
            chain.insert(1, via_step)
        for gi, g_items in enumerate(groups):
            panel = [
                {'kind': 'flow', 'title': 'Маълумот оқими',
                 'steps': chain,
                 **({'twoWay': True} if both else {})},
                {'kind': 'kpis', 'items': kpis},
                {'kind': 'list',
                 'title': 'ССВ тақдим этадиган маълумотлар' if outgoing else 'ССВга тақдим этиладиган маълумотлар',
                 'items': g_items},
            ]
            if both:
                panel.insert(3, {'kind': 'list', 'title': 'Идора ССВга тақдим этадиган маълумотлар',
                                 'items': [BIDIR[no]['in']]})
            if basis:
                panel.append({'kind': 'list', 'title': 'Ҳуқуқий асос', 'items': [basis]})
            agency_children[ag].append({
                'id': f'int-{no}' if len(groups) == 1 else f'int-{no}-{gi + 1}',
                'name': g_items[0],
                'label': trunc(g_items[0], 20),
                'mark': mark,
                'icon': sys_icon,
                'dir': 'both' if both else 'out' if outgoing else 'in',
                **({'via': True} if via else {}),
                'desc': f'{arrow_desc} Реестрдаги {no}-алмашинув. Идора тизими: {trunc(partner, 90)}',
                'status': 'live' if live else 'plan',
                'stat': {'value': f'№ {no}', 'label': 'реестр рақами'},
                'panel': panel,
            })

    # қўшимча йўналишлар (реестрдан ташқари)
    for ag, extras in EXTRA_LEAVES.items():
        for i, ex in enumerate(extras):
            e_short, e_mark, e_icon = moh_short(ex['moh_key'])
            e_out = ex['dir'] == 'out'
            e_live = ex['live']
            e_count_in_total = ex.get('count_in_total', True)
            if e_count_in_total:
                total += 1
                live_rows += 1 if e_live else 0
                out_rows += 1 if e_out else 0
            e_moh = {'title': e_short, 'sub': 'ССВ тизими', 'icon': e_icon}
            e_ag = {'title': AG[ag][1], 'sub': 'ҳамкор идора', 'icon': ag}
            e_basis = ex.get('basis', '')
            if e_basis:
                acts.add(e_basis)
            e_year = (YEAR_RE.search(e_basis) or [None, None])[1]
            e_kpis = [
                {'value': 'ССВ → Идора' if e_out else 'Идора → ССВ', 'label': 'маълумот йўналиши'},
                {'value': 'Мавжуд' if e_live else 'Режада', 'label': 'алмашинув ҳолати'},
            ]
            if not e_live:
                e_kpis.append({'value': '2026 йил охири', 'label': 'режа муддати'})
            if e_year:
                e_kpis.append({'value': e_year, 'label': 'асос йили'})
            e_panel = [
                {'kind': 'flow', 'title': 'Маълумот оқими',
                 'steps': [e_moh, e_ag] if e_out else [e_ag, e_moh]},
                {'kind': 'kpis', 'items': e_kpis},
                {'kind': 'list',
                 'title': 'ССВ тақдим этадиган маълумотлар' if e_out else 'ССВга тақдим этиладиган маълумотлар',
                 'items': [ex['name']]},
            ]
            if e_basis:
                e_panel.append({'kind': 'list', 'title': 'Ҳуқуқий асос', 'items': [e_basis]})
            e_registry_no = ex.get('registry_no')
            e_desc = (f'{e_short} → {AG[ag][0]}. ССВ маълумот тақдим этади.' if e_out
                      else f'{AG[ag][0]} → {e_short}. Идора ССВга маълумот тақдим этади.')
            if e_registry_no:
                e_desc += f' Реестрдаги {e_registry_no}-алмашинув.'
            if not e_live:
                e_desc += ' 2026 йил охирига режалаштирилган.'
            agency_children[ag].append({
                'id': f'int-x-{ag}-{i + 1}',
                'name': ex['name'],
                'label': trunc(ex['name'], 20),
                'mark': e_mark,
                'icon': e_icon,
                'dir': ex['dir'],
                'desc': e_desc,
                'status': 'live' if e_live else 'plan',
                'stat': {'value': f'№ {e_registry_no}' if e_registry_no else 'қўшимча',
                         'label': 'реестр рақами' if e_registry_no else 'йўналиш'},
                'panel': e_panel,
            })

    nodes = []
    for ag, (name, label, mark) in AG.items():
        if ag in EXCLUDE_AGENCIES:
            continue
        kids = agency_children[ag]
        if not kids:
            continue
        active = sum(1 for k in kids if k['status'] == 'live')
        planned = len(kids) - active
        out_n = sum(1 for k in kids if k['dir'] == 'out')
        in_n = len(kids) - out_n
        via_n = sum(1 for k in kids if k.get('via'))
        flows = ' · '.join(
            ([f'ССВ тақдим этади: {out_n} та'] if out_n else []) + ([f'ССВ қабул қилади: {in_n} та'] if in_n else []),
        )
        nodes.append({
            'id': f'ag-{ag}',
            'name': name,
            'label': label,
            'mark': mark,
            'icon': ag,
            **({'via': True} if via_n else {}),
            'desc': f'Алмашинув йўналишлари: {len(kids)} та, мавжуд: {active} та'
            + (f', режада: {planned} та (2026 йил охиригача)' if planned else '')
            + f'. {flows}.'
            + (f' {via_n} та йўналиш «Рақамли ҳукумат» платформаси орқали.' if via_n else ''),
            'stat': {'value': str(len(kids)), 'label': 'алмашинув йўналиши'},
            'children': kids,
        })

    header = (
        '/* ============================================================\n'
        '   АВТОГЕНЕРАЦИЯ — ССВ идоралараро алмашинув реестридан\n'
        f'   «Интеграция қилинадиган АТлар» (30.03.2026): {total} та йўналиш,\n'
        f'   {len(nodes)} та идора.\n'
        '   Қайта генерация: python3 scripts/gen-registry.py "<файл.xlsx>"\n'
        '   ============================================================ */\n'
        "import type { EcoNode } from '../lib/types'\n\n"
        f'export const REGISTRY_TOTALS = {{ total: {total}, active: {live_rows}, '
        f'process: {total - live_rows}, agencies: {len(nodes)}, acts: {len(acts)}, '
        f'outgoing: {out_rows}, incoming: {total - out_rows} }}\n\n'
        '/** Ҳамкор идоралар ва уларнинг интеграциялари (дарахт барглари) */\n'
        'export const GOV_AGENCY_NODES: EcoNode[] = '
    )
    out = header + json.dumps(nodes, ensure_ascii=False, indent=2) + '\n'
    with open('src/data/registry.ts', 'w', encoding='utf-8') as f:
        f.write(out)
    print(f'registry.ts: {total} rows, {len(nodes)} agencies, {len(acts)} acts, '
          f'live {live_rows}, out {out_rows}, via-platform {via_rows}')


if __name__ == '__main__':
    main(sys.argv[1])
